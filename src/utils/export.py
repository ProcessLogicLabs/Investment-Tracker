"""Excel export functionality for Asset Tracker."""

from typing import List, Dict, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from ..database.models import Asset


class ExcelExporter:
    """Export portfolio data to Excel format."""

    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.currency_format = '"$"#,##0.00'
        self.percent_format = '0.00%'
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def export(self, filename: str, assets: List[Asset], summary: Dict[str, Any]):
        """Export portfolio data to an Excel file."""
        wb = Workbook()

        # Create summary sheet
        self._create_summary_sheet(wb.active, summary)
        wb.active.title = "Summary"

        # Create assets sheet
        assets_sheet = wb.create_sheet("Assets")
        self._create_assets_sheet(assets_sheet, assets)

        # Create allocation sheet
        allocation_sheet = wb.create_sheet("Allocation")
        self._create_allocation_sheet(allocation_sheet, summary)

        wb.save(filename)

    def _create_summary_sheet(self, ws, summary: Dict[str, Any]):
        """Create the summary sheet."""
        # Title
        ws['A1'] = "Portfolio Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:C1')

        # Generated date
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(italic=True, color="666666")
        ws.merge_cells('A2:C2')

        # Summary data
        data = [
            ("", ""),
            ("Total Assets", summary.get('total_assets', 0)),
            ("Total Cost", summary.get('total_cost', 0)),
            ("Current Value", summary.get('total_value', 0)),
            ("Total Gain/Loss", summary.get('total_gain_loss', 0)),
            ("Return %", summary.get('gain_loss_percent', 0) / 100),
        ]

        row = 4
        for label, value in data:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            cell = ws.cell(row=row, column=2, value=value)

            if "Cost" in label or "Value" in label or "Gain/Loss" in label and "%" not in label:
                cell.number_format = self.currency_format
            elif "%" in label:
                cell.number_format = self.percent_format

            row += 1

        # Type breakdown
        row += 1
        ws.cell(row=row, column=1, value="Breakdown by Type").font = Font(bold=True, size=12)
        row += 1

        type_names = {
            'metal': 'Precious Metals',
            'stock': 'Securities',
            'realestate': 'Real Estate',
            'other': 'Other'
        }

        headers = ['Type', 'Count', 'Cost', 'Value', 'Gain/Loss']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        row += 1

        by_type = summary.get('by_type', {})
        for asset_type, data in by_type.items():
            ws.cell(row=row, column=1, value=type_names.get(asset_type, asset_type))
            ws.cell(row=row, column=2, value=data.get('count', 0))

            cost_cell = ws.cell(row=row, column=3, value=data.get('total_cost', 0))
            cost_cell.number_format = self.currency_format

            value_cell = ws.cell(row=row, column=4, value=data.get('current_value', 0))
            value_cell.number_format = self.currency_format

            gain_loss = data.get('current_value', 0) - data.get('total_cost', 0)
            gl_cell = ws.cell(row=row, column=5, value=gain_loss)
            gl_cell.number_format = self.currency_format

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

    def _create_assets_sheet(self, ws, assets: List[Asset]):
        """Create the assets detail sheet."""
        headers = [
            'Name', 'Type', 'Symbol', 'Quantity', 'Unit', 'Wt/Unit', 'Total Wt',
            'Purchase Price', 'Current Price', 'Total Cost', 'Current Value',
            'Gain/Loss', 'Gain/Loss %', 'Purchase Date', 'Notes'
        ]

        # Header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

        type_names = {
            'metal': 'Precious Metal',
            'stock': 'Stock/Security',
            'realestate': 'Real Estate',
            'other': 'Other'
        }

        # Data rows
        for row, asset in enumerate(assets, 2):
            ws.cell(row=row, column=1, value=asset.name)
            ws.cell(row=row, column=2, value=type_names.get(asset.asset_type, asset.asset_type))
            ws.cell(row=row, column=3, value=asset.symbol)
            ws.cell(row=row, column=4, value=asset.quantity)
            ws.cell(row=row, column=5, value=asset.unit or '')
            ws.cell(row=row, column=6, value=asset.weight_per_unit if asset.asset_type == 'metal' else '')
            ws.cell(row=row, column=7, value=asset.total_weight if asset.asset_type == 'metal' else '')

            pp_cell = ws.cell(row=row, column=8, value=asset.purchase_price)
            pp_cell.number_format = self.currency_format

            cp_cell = ws.cell(row=row, column=9, value=asset.current_price)
            cp_cell.number_format = self.currency_format

            tc_cell = ws.cell(row=row, column=10, value=asset.total_cost)
            tc_cell.number_format = self.currency_format

            cv_cell = ws.cell(row=row, column=11, value=asset.current_value)
            cv_cell.number_format = self.currency_format

            gl_cell = ws.cell(row=row, column=12, value=asset.gain_loss)
            gl_cell.number_format = self.currency_format

            glp_cell = ws.cell(row=row, column=13, value=asset.gain_loss_percent / 100)
            glp_cell.number_format = self.percent_format

            ws.cell(row=row, column=14, value=asset.purchase_date or '')
            ws.cell(row=row, column=15, value=asset.notes or '')

            # Apply borders
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = self.thin_border

        # Adjust column widths
        column_widths = [20, 15, 12, 12, 8, 8, 10, 15, 15, 15, 15, 15, 12, 12, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Add totals row
        if assets:
            total_row = len(assets) + 2
            ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)

            total_cost = sum(a.total_cost for a in assets)
            total_value = sum(a.current_value for a in assets)
            total_gain_loss = total_value - total_cost

            tc_cell = ws.cell(row=total_row, column=10, value=total_cost)
            tc_cell.number_format = self.currency_format
            tc_cell.font = Font(bold=True)

            cv_cell = ws.cell(row=total_row, column=11, value=total_value)
            cv_cell.number_format = self.currency_format
            cv_cell.font = Font(bold=True)

            gl_cell = ws.cell(row=total_row, column=12, value=total_gain_loss)
            gl_cell.number_format = self.currency_format
            gl_cell.font = Font(bold=True)

    def _create_allocation_sheet(self, ws, summary: Dict[str, Any]):
        """Create the allocation breakdown sheet."""
        ws['A1'] = "Asset Allocation"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')

        headers = ['Asset Type', 'Value', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment

        type_names = {
            'metal': 'Precious Metals',
            'stock': 'Securities',
            'realestate': 'Real Estate',
            'other': 'Other'
        }

        total_value = summary.get('total_value', 0)
        by_type = summary.get('by_type', {})

        row = 4
        for asset_type, data in by_type.items():
            value = data.get('current_value', 0)
            percentage = (value / total_value) if total_value > 0 else 0

            ws.cell(row=row, column=1, value=type_names.get(asset_type, asset_type))

            value_cell = ws.cell(row=row, column=2, value=value)
            value_cell.number_format = self.currency_format

            pct_cell = ws.cell(row=row, column=3, value=percentage)
            pct_cell.number_format = self.percent_format

            row += 1

        # Total row
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        total_cell = ws.cell(row=row, column=2, value=total_value)
        total_cell.number_format = self.currency_format
        total_cell.font = Font(bold=True)

        pct_cell = ws.cell(row=row, column=3, value=1)
        pct_cell.number_format = self.percent_format
        pct_cell.font = Font(bold=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
